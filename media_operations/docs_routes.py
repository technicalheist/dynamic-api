import os
import uuid
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook

docs_routes = Blueprint("docs_routes", __name__)

UPLOAD_FOLDER = "uploads/documents"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def generate_random_id():
    return str(uuid.uuid4())

@docs_routes.route("/docs", methods=["POST"])
def upload_and_process_excel():
    if "file" not in request.files:
        return jsonify({
            "status": "ERROR",
            "message": "No file part",
            "error": []
        }), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({
            "status": "ERROR",
            "message": "No selected file",
            "error": []
        }), 400

    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        return jsonify({
            "status": "ERROR",
            "message": "Only Excel files are allowed",
            "error": []
        }), 400

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    try:
        file.save(file_path)
    except Exception as e:
        return jsonify({
            "status": "ERROR",
            "message": "Error saving file",
            "error": [str(e)]
        }), 500

    try:
        # Load the Excel file
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Read the first sheet
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return jsonify({
                "status": "ERROR",
                "message": "Excel sheet is empty",
                "error": []
            }), 400

        # Extract header and validate
        header = rows[0]
        required_columns = {"firstName", "lastName", "emailAddress"}
        if not required_columns.issubset(set(header)):
            return jsonify({
                "status": "ERROR",
                "message": "Missing required columns",
                "error": []
            }), 400

        # Map rows to objects
        data = []
        for row in rows[1:]:
            row_data = dict(zip(header, row))
            data.append({
                "id": generate_random_id(),
                "firstName": row_data.get("firstName", ""),
                "lastName": row_data.get("lastName", ""),
                "emailAddress": row_data.get("emailAddress", ""),
                "phoneNumber": row_data.get("phoneNumber", ""),
                "isAdmin": False,
                "isActive": False,
                "isDeleted": False
            })

        return jsonify({
            "status": "SUCCESS",
            "message": "SUCCESS",
            "data": data
        }), 200

    except Exception as e:
        return jsonify({
            "status": "ERROR",
            "message": "Error processing Excel file",
            "error": [str(e)]
        }), 500
